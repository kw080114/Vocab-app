import json
from pathlib import Path

import openpyxl

workbook_path = Path(r"C:\Users\kw080\Downloads\studytoday_sat_vocab_500.xlsx")
output_path = Path.cwd() / "data" / "studytoday-sat-vocab.json"
set_name = "StudyToday SAT Vocab"

workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
worksheet = workbook[workbook.sheetnames[0]]
rows = list(worksheet.iter_rows(values_only=True))
headers = [str(value).strip() for value in rows[0]]

word_index = headers.index("word")
part_of_speech_index = headers.index("partOfSpeech")
definition_index = headers.index("definition")

entries = []

for source_order, row in enumerate(rows[1:], start=1):
    word = row[word_index]
    part_of_speech = row[part_of_speech_index]
    definition = row[definition_index]

    if not word or not definition:
        continue

    entries.append(
        {
            "sourceOrder": source_order,
            "setName": set_name,
            "word": str(word).strip(),
            "definition": str(definition).strip(),
            "example": "",
            "partOfSpeech": str(part_of_speech or "").strip(),
            "difficulty": 2,
            "tags": ["SAT", "StudyToday"],
        }
    )

output_path.parent.mkdir(parents=True, exist_ok=True)
output_path.write_text(json.dumps(entries, indent=2) + "\n", encoding="utf-8")

print(f"Extracted {len(entries)} StudyToday vocabulary entries.")
print(f"Wrote {output_path}")
